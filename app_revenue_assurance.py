import os
import json
import datetime
import re
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
import plotly.express as px

# 1. Configuração Inicial da Página
st.set_page_config(
    page_title="Grupo Arbaitman | Revenue Assurance",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded"
)

ARQUIVO_DASHBOARD = "Dashboard_Revenue_Assurance_Consolidado.xlsx"
ARQUIVO_NAO_CONCILIADOS = "NÃO CONCILIADOS.xlsx"
ARQUIVO_CONCILIADOS = "CONCILIADOS _ REGULARIZADOS.xlsx"
ARQUIVO_USUARIOS = "usuarios_autorizados.json"

USUARIOS_PADRAO = {
    "mribeiro": {"senha": "123", "nome": "Marcos Ribeiro", "perfil": "Compliance", "status": "APROVADO", "data_solicitacao": "2026-08-31"},
    "compliance1": {"senha": "123", "nome": "Compliance - Auditoria 01", "perfil": "Compliance", "status": "APROVADO", "data_solicitacao": "2026-08-31"},
    "operacao": {"senha": "123", "nome": "Equipe Operacional", "perfil": "Operacao", "status": "APROVADO", "data_solicitacao": "2026-08-31"},
    "backoffice": {"senha": "123", "nome": "Atendimento Backoffice", "perfil": "Operacao", "status": "APROVADO", "data_solicitacao": "2026-08-31"}
}

def carregar_usuarios():
    if not os.path.exists(ARQUIVO_USUARIOS):
        with open(ARQUIVO_USUARIOS, "w", encoding="utf-8") as f:
            json.dump(USUARIOS_PADRAO, f, ensure_ascii=False, indent=4)
        return USUARIOS_PADRAO
    try:
        with open(ARQUIVO_USUARIOS, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return USUARIOS_PADRAO

def salvar_usuarios(dict_users):
    with open(ARQUIVO_USUARIOS, "w", encoding="utf-8") as f:
        json.dump(dict_users, f, ensure_ascii=False, indent=4)

usuarios_db = carregar_usuarios()

# 2. Estilização CSS Corporativa
st.markdown("""
    <style>
        .stApp { background-color: #f8f9fa; }
        
        div.stButton > button[kind="primary"], div.stButton > button {
            background-color: #002060 !important;
            color: #ffffff !important;
            border-radius: 6px !important;
            border: none !important;
            font-weight: 600 !important;
        }
        div.stButton > button:hover {
            background-color: #001040 !important;
            color: #ffffff !important;
        }
        
        .header-box {
            background: linear-gradient(135deg, #002060 0%, #003366 100%);
            padding: 15px 25px;
            border-radius: 10px;
            color: white;
            margin-bottom: 20px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.06);
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
        .header-box h1 { color: #ffffff !important; margin: 0; font-size: 24px; font-weight: 700; }
        .header-box p { color: #d0e0ff !important; margin-top: 4px; font-size: 13px; margin-bottom: 0; }
        
        .login-card {
            background-color: #ffffff;
            padding: 30px 35px;
            border-radius: 12px;
            border-top: 6px solid #002060;
            box-shadow: 0 8px 24px rgba(0,0,0,0.08);
            margin-top: 20px;
        }
        
        div[data-testid="stMetric"] {
            background-color: #ffffff;
            border-radius: 8px;
            padding: 14px;
            border-left: 5px solid #002060;
            box-shadow: 0 2px 6px rgba(0,0,0,0.04);
        }

        [data-baseweb="tag"], 
        span[data-baseweb="tag"], 
        div[data-baseweb="tag"], 
        li[data-baseweb="tag"] {
            background-color: #6c757d !important;
            background: #6c757d !important;
            color: #ffffff !important;
            border-radius: 4px !important;
        }
        
        [data-baseweb="tag"] span,
        span[data-baseweb="tag"] span {
            color: #ffffff !important;
        }

        .brand-header {
            font-size: 22px;
            font-weight: 800;
            color: #002060;
            letter-spacing: 1px;
            text-align: center;
            margin-bottom: 5px;
        }
    </style>
""", unsafe_allow_html=True)

def renderizar_marca():
    st.markdown('<div class="brand-header">GRUPO ARBAITMAN</div>', unsafe_allow_html=True)

def gerar_excel_formatado(df_export, nome_aba="Relatorio_Filtrado"):
    buffer = io.BytesIO()
    
    if df_export is None or df_export.empty:
        df_export = pd.DataFrame(columns=["Aviso"], data=[["Nenhum registro encontrado para os filtros selecionados"]])

    cols_remover = [c for c in ["Dt_Parsed", "Mes_Ano_Sort", "Bilhetes_Str"] if c in df_export.columns]
    df_clean = df_export.drop(columns=cols_remover) if cols_remover else df_export

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_clean.to_excel(writer, sheet_name=nome_aba[:30], index=False)
        
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active
    
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    currency_cols = ['A vista', 'A credito', 'Taxa', 'Comissão', 'Taxa DU', 'Desc.', 'Incentivo', 'VL. Líquido']

    ws.views.sheetView[0].showGridLines = True

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    max_row = ws.max_row
    max_col = ws.max_column

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        col_name = str(ws.cell(row=1, column=col_idx).value or '')
        
        len_vals = [len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, max_row + 1)]
        max_len = max(len_vals) if len_vals else 10
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
        for r in range(2, max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = thin_border
            
            if col_name in currency_cols:
                try:
                    if cell.value is not None and str(cell.value).strip() not in ['-', '']:
                        cell.value = float(cell.value)
                        cell.number_format = 'R$ #,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                except:
                    pass

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# 3. Autenticação e Sessão Persistente
if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False
if "usuario_atual" not in st.session_state:
    st.session_state["usuario_atual"] = None
if "perfil_atual" not in st.session_state:
    st.session_state["perfil_atual"] = None
if "login_user_id" not in st.session_state:
    st.session_state["login_user_id"] = None

if not st.session_state["autenticado"]:
    col_left, col_center, col_right = st.columns([1, 1.2, 1])
    
    with col_center:
        st.markdown('<div class="login-card">', unsafe_allow_html=True)
        renderizar_marca()
        st.markdown("<p style='text-align: center; color: #6c757d; font-size: 13px; margin-bottom: 20px;'>Maringá Turismo | Portal de Revenue Assurance</p>", unsafe_allow_html=True)
        
        aba_login, aba_redefinir, aba_solicitar = st.tabs(["🔐 Entrar", "🔑 Esqueci a Senha", "📝 Solicitar Acesso"])
        
        with aba_login:
            user_input = st.text_input("Usuário de Acesso:", key="login_user").strip().lower()
            pass_input = st.text_input("Senha:", type="password", key="login_pass").strip()
            btn_entrar = st.button("🚀 Acessar Sistema", type="primary")
            
            if btn_entrar:
                if user_input in usuarios_db:
                    dados_u = usuarios_db[user_input]
                    if dados_u["senha"] == pass_input:
                        if dados_u.get("status") == "APROVADO":
                            st.session_state["autenticado"] = True
                            st.session_state["usuario_atual"] = dados_u["nome"]
                            st.session_state["perfil_atual"] = dados_u["perfil"]
                            st.session_state["login_user_id"] = user_input
                            st.rerun()
                        elif dados_u.get("status") == "PENDENTE":
                            st.warning("⏳ Sua solicitação de acesso está **Pendente de Aprovação** pelo Compliance.")
                        else:
                            st.error("❌ Acesso não autorizado.")
                    else:
                        st.error("❌ Senha incorreta.")
                else:
                    st.error("❌ Usuário não cadastrado.")

        with aba_redefinir:
            st.caption("Redefinição direta de senha de acesso.")
            with st.form("form_redefinir_senha_login"):
                user_reset = st.text_input("Informe seu Usuário de Acesso:").strip().lower()
                nova_senha_login = st.text_input("Nova Senha:", type="password")
                confirma_senha_login = st.text_input("Confirme a Nova Senha:", type="password")
                btn_redefinir_senha = st.form_submit_button("🔄 Redefinir Senha")

                if btn_redefinir_senha:
                    if not user_reset:
                        st.error("⚠️ Digite o usuário de acesso.")
                    elif user_reset not in usuarios_db:
                        st.error("❌ Usuário não encontrado no sistema.")
                    elif not nova_senha_login.strip():
                        st.error("⚠️ Digite a nova senha.")
                    elif nova_senha_login != confirma_senha_login:
                        st.error("❌ As senhas não coincidem.")
                    else:
                        usuarios_db[user_reset]["senha"] = nova_senha_login.strip()
                        salvar_usuarios(usuarios_db)
                        st.success("✅ Senha redefinida com sucesso! Clique na aba '🔐 Entrar' para acessar.")

        with aba_solicitar:
            st.caption("Solicitação formal de acesso para novos colaboradores.")
            novo_nome = st.text_input("Nome Completo:")
            novo_user = st.text_input("Usuário Desejado (ex: nome.sobrenome):").strip().lower()
            nova_senha = st.text_input("Crie uma Senha:", type="password")
            novo_perfil = st.selectbox("Perfil Solicitado:", options=["Operacao", "Compliance"])
            
            st.markdown("""
                <div style="background-color: #f8f9fa; border: 1px solid #dee2e6; padding: 10px; border-radius: 6px; font-size: 11px; color: #495057; max-height: 90px; overflow-y: scroll; margin-bottom: 10px;">
                    <b>TERMO DE CONFIDENCIALIDADE - GRUPO ARBAITMAN</b><br>
                    O usuário declara estar ciente do caráter confidencial das informações de Revenue Assurance. Qualquer alteração ou exportação de dados é monitorada na trilha de auditoria.
                </div>
            """, unsafe_allow_html=True)
            
            termo_aceito = st.checkbox("Li e aceito os termos de sigilo")
            btn_cadastrar = st.button("📩 Enviar Solicitação")
            
            if btn_cadastrar:
                if not novo_nome or not novo_user or not nova_senha:
                    st.error("⚠️ Preencha todos os campos do formulário.")
                elif not termo_aceito:
                    st.error("⚠️ Aceite o termo de sigilo para continuar.")
                elif novo_user in usuarios_db:
                    st.error("❌ Usuário já existente.")
                else:
                    usuarios_db[novo_user] = {
                        "senha": nova_senha,
                        "nome": novo_nome,
                        "perfil": novo_perfil,
                        "status": "PENDENTE",
                        "data_solicitacao": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    salvar_usuarios(usuarios_db)
                    st.success("✅ Solicitação enviada! Aguarde a liberação pelo Compliance.")

        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# 4. Funções de Padronização, Limpeza de Linhas TOTAL e Tratamento
def padronizar_gerentes_e_setores(val):
    if pd.isna(val) or not str(val).strip():
        return "Não Atribuído"
    v = str(val).strip()
    
    if v in ["Rosangela Pallu", "Rosângela Pallu", "ROSANGELA PALLU", "ROSÂNGELA PALLU"]:
        return "Rosângela Pallu"
    elif v in ["CENTRAL DE EVENTOS", "Central de Eventos", "central de eventos"]:
        return "Central de Eventos"
    elif v in ["Jaime Schinaider", "Jaime Schnaider", "JAIME SCHNAIDER"]:
        return "Jaime Schnaider"
    elif v in ["Silvana Celane", "Silvana Celani", "SILVANA CELANI"]:
        return "Silvana Celani"
    elif "suporte" in v.lower() or "benner" in v.lower() or "katia" in v.lower():
        return "Suporte Backoffice"
    return v

def categorizar_tipo_inconsistencia(row):
    origem = str(row.get('Origem_Aba', ''))
    status_div = str(row.get('Status_Divergencia', '')).strip()
    status_sis = str(row.get('Status_Sistema', '')).strip()
    
    if origem == "99_Geral" or status_sis == "NAO_CONSTA" or "Pendente" in str(row.get('Status_Geral', '')):
        return "Pendente de Lançamento no ERP"
    elif status_div and status_div not in ["nan", "Valores Corretos", ""]:
        return status_div
    elif "Sem_Divergencia" in origem or status_div == "Valores Corretos":
        return "Sem Divergência (Conciliado)"
    else:
        return "Pendente de Lançamento no ERP"

def padronizar_e_deduplicar_colunas(df, origem=""):
    if df is None or df.empty: return pd.DataFrame()
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # REMOÇÃO SEGURA DAS LINHAS DE TOTALIZAÇÃO EXCEL
    col_pv = 'Ponto de venda' if 'Ponto de venda' in df.columns else None
    col_bil = 'Bilhetes' if 'Bilhetes' in df.columns else ('Bilhete' if 'Bilhete' in df.columns else None)

    mask_descarte = pd.Series(False, index=df.index)
    if col_pv:
        mask_descarte = mask_descarte | df[col_pv].astype(str).str.strip().str.upper().str.contains("TOTAL", na=False)
    if col_bil:
        mask_descarte = mask_descarte | df[col_bil].astype(str).str.strip().str.upper().str.contains("TOTAL", na=False) | df[col_bil].isna()

    df = df[~mask_descarte].copy()

    col_area_resp = None
    renomear = {}
    for col in df.columns:
        c_str = str(col).strip().lower()
        if c_str in ["área responsável", "area responsavel"]:
            col_area_resp = col
            renomear[col] = "Área Resp. Operação"
        elif c_str in ["gerente", "gerentes", "área resp. operação", "area resp. operacao", "area_resp_operacao"]:
            renomear[col] = "Área Resp. Operação"
        elif c_str in ["obs", "obs. operação", "observação", "observacao"]:
            renomear[col] = "Obs. Operação"
        elif c_str == "setor":
            renomear[col] = "Setor"
        elif c_str in ["consultor_lemon", "emissor", "consultor"]:
            renomear[col] = "Consultor_Lemon"
        elif c_str in ["bilhete", "bilhetes"]:
            renomear[col] = "Bilhetes"
        elif c_str in ["data emissão", "data emissao", "data_emissao"]:
            renomear[col] = "Data Emissão"
            
    df_out = df.rename(columns=renomear).copy()
    df_out = df_out.loc[:, ~df_out.columns.duplicated()].copy()
    
    if "Área Resp. Operação" not in df_out.columns: df_out["Área Resp. Operação"] = "Não Atribuído"
    if "Setor" not in df_out.columns: df_out["Setor"] = "Geral"
    if "Consultor_Lemon" not in df_out.columns: df_out["Consultor_Lemon"] = "-"
    if "Localizador_Sistema" not in df_out.columns: df_out["Localizador_Sistema"] = "-"
    if "Status_Geral" not in df_out.columns: df_out["Status_Geral"] = "Pendente de Lançamento"
    if "Obs. Operação" not in df_out.columns: df_out["Obs. Operação"] = ""
    if "Data Emissão" not in df_out.columns: df_out["Data Emissão"] = "-"

    df_out["Área Resp. Operação"] = df_out["Área Resp. Operação"].apply(padronizar_gerentes_e_setores)
    df_out["Setor"] = df_out["Setor"].apply(lambda s: "Central de Eventos" if str(s).upper() == "CENTRAL DE EVENTOS" else s)

    val_area_orig = df[col_area_resp].astype(str).str.lower() if col_area_resp and col_area_resp in df.columns else pd.Series("", index=df.index)
    val_ger_orig = df_out["Área Resp. Operação"].astype(str).str.lower()
    val_obs_orig = df_out["Obs. Operação"].fillna("").astype(str).str.lower()

    mascara_suporte = (
        val_area_orig.str.contains("suporte benner|suporte backoffice|suporte ti|backoffice", na=False) |
        val_ger_orig.str.contains("suporte benner|suporte backoffice|katia martins", na=False) |
        val_obs_orig.str.contains("ticket 375|chamado no backoffice|chamado backoffice|erro de integração|erro integração|não integrou|não integração|aguardando suporte", na=False)
    )
    
    df_out.loc[mascara_suporte, "Área Resp. Operação"] = "Suporte Backoffice"
    df_out["Origem_Aba"] = origem
    df_out["Tipo_Inconsistencia"] = df_out.apply(categorizar_tipo_inconsistencia, axis=1)

    df_out["Dt_Parsed"] = pd.to_datetime(df_out["Data Emissão"], errors="coerce", dayfirst=True)
    df_out["Mes_Ano_Sort"] = df_out["Dt_Parsed"].dt.strftime("%Y-%m")
    
    mapa_meses = {"01": "Jan", "02": "Fev", "03": "Mar", "04": "Abr", "05": "Mai", "06": "Jun", "07": "Jul", "08": "Ago", "09": "Set", "10": "Out", "11": "Nov", "12": "Dez"}
    
    def extrair_rotulo_mes(row):
        if pd.notna(row["Dt_Parsed"]):
            m_num = row["Dt_Parsed"].strftime("%m")
            y_num = row["Dt_Parsed"].strftime("%Y")
            return f"{mapa_meses.get(m_num, m_num)}/{y_num}"
        s_raw = str(row["Data Emissão"])
        m_match = re.search(r'/(\d{2})/(\d{4})', s_raw)
        if m_match:
            m_num, y_num = m_match.group(1), m_match.group(2)
            return f"{mapa_meses.get(m_num, m_num)}/{y_num}"
        return "Acumulado / Sem Data"

    df_out["Mes_Ano_Label"] = df_out.apply(extrair_rotulo_mes, axis=1)
    return df_out

@st.cache_data(ttl=2)
def carregar_bases():
    if not os.path.exists(ARQUIVO_DASHBOARD):
        return None, None, None, None, None, "ARQUIVO_NAO_ENCONTRADO"
    try:
        xls = pd.ExcelFile(ARQUIVO_DASHBOARD)
        aba_names = xls.sheet_names
        
        df_m = pd.read_excel(xls, sheet_name="99_Base_Divergencias_Geral") if "99_Base_Divergencias_Geral" in aba_names else pd.DataFrame()
        df_d = pd.read_excel(xls, sheet_name="98_OK_Divergencia_Operacao") if "98_OK_Divergencia_Operacao" in aba_names else pd.DataFrame()
        df_s = pd.read_excel(xls, sheet_name="98_OK_Sem_Divergencia_Concil") if "98_OK_Sem_Divergencia_Concil" in aba_names else pd.DataFrame()
        df_b = pd.read_excel(xls, sheet_name="99_Suporte backoffice") if "99_Suporte backoffice" in aba_names else pd.DataFrame()
        
        df_master = padronizar_e_deduplicar_colunas(df_m, "99_Geral")
        df_div_op = padronizar_e_deduplicar_colunas(df_d, "98_Divergencia_Operacao")
        df_sem_div = padronizar_e_deduplicar_colunas(df_s, "98_Sem_Divergencia")
        df_backoffice = padronizar_e_deduplicar_colunas(df_b, "99_Backoffice")
        
        if "00_Log_Auditoria" in aba_names:
            df_log = pd.read_excel(xls, sheet_name="00_Log_Auditoria")
            df_log = df_log.loc[:, ~df_log.columns.duplicated()].copy()
        else:
            df_log = pd.DataFrame(columns=["Data_Hora", "Bilhete", "Usuario_Acao", "Status_Anterior", "Novo_Status", "Area_Anterior", "Nova_Area", "Observacao", "Tipo_Interacao"])
            
        return df_master, df_div_op, df_sem_div, df_backoffice, df_log, "OK"
    except PermissionError:
        return None, None, None, None, None, "ARQUIVO_BLOQUEADO"
    except Exception as e:
        return None, None, None, None, None, str(e)

def sincronizar_planilhas_auxiliares(bilhete_str, nova_area_resp, observacao_str):
    bilhete_target = str(bilhete_str).strip()
    area_gravar = "SUPORTE BENNER" if str(nova_area_resp).lower().strip() in ["suporte backoffice", "suporte benner"] else nova_area_resp

    if os.path.exists(ARQUIVO_NAO_CONCILIADOS):
        try:
            df_nc = pd.read_excel(ARQUIVO_NAO_CONCILIADOS, sheet_name="NÃO CONCILIADOS")
            col_b = "BILHETES" if "BILHETES" in df_nc.columns else ("Bilhete" if "Bilhete" in df_nc.columns else None)
            if col_b:
                m_nc = df_nc[col_b].astype(str).str.strip() == bilhete_target
                if m_nc.any():
                    if "ÁREA RESPONSÁVEL" in df_nc.columns: df_nc.loc[m_nc, "ÁREA RESPONSÁVEL"] = area_gravar
                    elif "GERENTES" in df_nc.columns: df_nc.loc[m_nc, "GERENTES"] = area_gravar
                    if "OBS" in df_nc.columns: df_nc.loc[m_nc, "OBS"] = observacao_str
                    with pd.ExcelWriter(ARQUIVO_NAO_CONCILIADOS, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_nc.to_excel(writer, sheet_name="NÃO CONCILIADOS", index=False)
        except Exception as e:
            st.warning(f"⚠️ Nota ao atualizar 'NÃO CONCILIADOS.xlsx': {str(e)}")

    if os.path.exists(ARQUIVO_CONCILIADOS):
        try:
            xls_c = pd.ExcelFile(ARQUIVO_CONCILIADOS)
            sheet_target = xls_c.sheet_names[0]
            df_cr = pd.read_excel(xls_c, sheet_name=sheet_target)
            col_b = "BILHETES" if "BILHETES" in df_cr.columns else ("Bilhete" if "Bilhete" in df_cr.columns else None)
            if col_b:
                m_cr = df_cr[col_b].astype(str).str.strip() == bilhete_target
                if m_cr.any():
                    if "ÁREA RESPONSÁVEL" in df_cr.columns: df_cr.loc[m_cr, "ÁREA RESPONSÁVEL"] = area_gravar
                    elif "GERENTES" in df_cr.columns: df_cr.loc[m_cr, "GERENTES"] = area_gravar
                    if "OBS" in df_cr.columns: df_cr.loc[m_cr, "OBS"] = observacao_str
                    with pd.ExcelWriter(ARQUIVO_CONCILIADOS, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_cr.to_excel(writer, sheet_name=sheet_target, index=False)
        except Exception as e:
            st.warning(f"⚠️ Nota ao atualizar 'CONCILIADOS _ REGULARIZADOS.xlsx': {str(e)}")

df_master, df_div_op, df_sem_div, df_backoffice, df_log_master, status_carga = carregar_bases()

if status_carga == "ARQUIVO_BLOQUEADO":
    st.error(f"⚠️ O arquivo **'{ARQUIVO_DASHBOARD}'** está em uso. Feche a planilha para atualizar.")
    if st.button("🔄 Recarregar Dados"):
        st.cache_data.clear()
        st.rerun()
    st.stop()

if (df_master is None or df_master.empty) and (df_div_op is None or df_div_op.empty):
    st.error("⚠️ Base consolidada de divergências não encontrada.")
    st.stop()

COL_GERENTE = "Área Resp. Operação"
COL_SETOR = "Setor"
COL_EMISSOR = "Consultor_Lemon"

# CONSOLIDAÇÃO TOTAL (415 EMISSÕES SEM A LINHA TOTAL)
df_list = [d for d in [df_master, df_div_op, df_sem_div] if not d.empty]
df_acao_total = pd.concat(df_list, ignore_index=True, sort=False) if len(df_list) > 0 else pd.DataFrame()

if "Bilhetes" in df_acao_total.columns:
    df_acao_total["Bilhetes_Str"] = df_acao_total["Bilhetes"].astype(str).str.strip()

usuario_log_formatado = f"{st.session_state['usuario_atual']} ({st.session_state['login_user_id']})"

col_hdr1, col_hdr2 = st.columns([3, 1])
with col_hdr1:
    st.markdown(f"""
        <div class="header-box">
            <div>
                <h1>GRUPO ARBAITMAN | Revenue Assurance</h1>
                <p>Maringá Turismo — Conectado como: <b>{st.session_state['usuario_atual']}</b> | Perfil: <b>{st.session_state['perfil_atual']}</b></p>
            </div>
        </div>
    """, unsafe_allow_html=True)

with col_hdr2:
    if st.button("🔄 Recarregar / Atualizar Dados", type="primary"):
        st.cache_data.clear()
        st.rerun()

if "msg_sucesso" in st.session_state:
    st.success(st.session_state["msg_sucesso"])
    del st.session_state["msg_sucesso"]

renderizar_marca()
st.sidebar.markdown(f"### 👤 {st.session_state['usuario_atual']}")

col_btn_sair, col_btn_senha = st.sidebar.columns(2)
with col_btn_sair:
    if st.button("🔒 Sair"):
        st.session_state["autenticado"] = False
        st.session_state["usuario_atual"] = None
        st.session_state["perfil_atual"] = None
        st.rerun()

with col_btn_senha:
    btn_mudar_senha = st.button("🔑 Senha")

if btn_mudar_senha or st.session_state.get("abrir_modal_senha", False):
    st.session_state["abrir_modal_senha"] = True
    with st.sidebar.expander("🔑 Alterar Minha Senha", expanded=True):
        u_id_atual = st.session_state["login_user_id"]
        with st.form("form_senha_sidebar"):
            senha_atual_input = st.text_input("Senha Atual:", type="password")
            nova_senha_input = st.text_input("Nova Senha:", type="password")
            confirma_senha_input = st.text_input("Confirmar Nova Senha:", type="password")
            btn_upd_pwd = st.form_submit_button("💾 Atualizar Senha")
            
            if btn_upd_pwd:
                if usuarios_db[u_id_atual]["senha"] != senha_atual_input:
                    st.error("❌ Senha atual incorreta.")
                elif not nova_senha_input.strip():
                    st.error("⚠️ Digite uma nova senha.")
                elif nova_senha_input != confirma_senha_input:
                    st.error("❌ As senhas não coincidem.")
                else:
                    usuarios_db[u_id_atual]["senha"] = nova_senha_input.strip()
                    salvar_usuarios(usuarios_db)
                    st.session_state["msg_sucesso"] = "✅ Senha alterada com sucesso!"
                    st.session_state["abrir_modal_senha"] = False
                    st.rerun()

st.sidebar.markdown("---")
st.sidebar.title("🔍 Filtros Operacionais")

# FILTROS DINÂMICOS
df_meses = df_acao_total.dropna(subset=["Mes_Ano_Label"]).sort_values(by="Mes_Ano_Sort")
opcoes_meses = sorted(list(df_meses["Mes_Ano_Label"].unique()))
mes_sel = st.sidebar.multiselect("📅 Mês de Emissão:", options=opcoes_meses, default=[])

df_f_mes = df_acao_total[df_acao_total["Mes_Ano_Label"].isin(mes_sel)] if len(mes_sel) > 0 else df_acao_total

opcoes_setor = sorted(list(df_f_mes[COL_SETOR].dropna().astype(str).unique()))
setor_sel = st.sidebar.multiselect("Setor Responsável:", options=opcoes_setor, default=[])

df_f_1 = df_f_mes[df_f_mes[COL_SETOR].astype(str).isin(setor_sel)] if len(setor_sel) > 0 else df_f_mes

opcoes_gerente = sorted(list(df_f_1[COL_GERENTE].dropna().astype(str).unique()))
gerente_sel = st.sidebar.multiselect("Gerente (Área Resp. Operação):", options=opcoes_gerente, default=[])

df_f_2 = df_f_1[df_f_1[COL_GERENTE].astype(str).isin(gerente_sel)] if len(gerente_sel) > 0 else df_f_1

opcoes_emissor = sorted(list(df_f_2[COL_EMISSOR].dropna().astype(str).unique()))
emissor_sel = st.sidebar.multiselect("Emissor / Consultor (OBT):", options=opcoes_emissor, default=[])

df_f_3 = df_f_2[df_f_2[COL_EMISSOR].astype(str).isin(emissor_sel)] if len(emissor_sel) > 0 else df_f_2

tipos_emissao = sorted(list(df_f_3["Tipo_Emissao_Lemon"].dropna().astype(str).unique())) if "Tipo_Emissao_Lemon" in df_f_3.columns else []
emissao_sel = st.sidebar.multiselect("Tipo de Emissão (OBT):", options=tipos_emissao, default=[])

df_f_4 = df_f_3[df_f_3["Tipo_Emissao_Lemon"].astype(str).isin(emissao_sel)] if len(emissao_sel) > 0 and "Tipo_Emissao_Lemon" in df_f_3.columns else df_f_3

cias = sorted(list(df_f_4["CIA"].dropna().astype(str).unique())) if "CIA" in df_f_4.columns else []
cia_sel = st.sidebar.multiselect("Companhia Aérea:", options=cias, default=[])

def aplicar_filtros_globais(df):
    if df is None or df.empty: return df
    m = pd.Series(True, index=df.index)
    if "Mes_Ano_Label" in df.columns and len(mes_sel) > 0: m = m & df["Mes_Ano_Label"].isin(mes_sel)
    if COL_SETOR in df.columns and len(setor_sel) > 0: m = m & df[COL_SETOR].astype(str).isin(setor_sel)
    if COL_GERENTE in df.columns and len(gerente_sel) > 0: m = m & df[COL_GERENTE].astype(str).isin(gerente_sel)
    if COL_EMISSOR in df.columns and len(emissor_sel) > 0: m = m & df[COL_EMISSOR].astype(str).isin(emissor_sel)
    if "Tipo_Emissao_Lemon" in df.columns and len(emissao_sel) > 0: m = m & df["Tipo_Emissao_Lemon"].astype(str).isin(emissao_sel)
    if "CIA" in df.columns and len(cia_sel) > 0: m = m & df["CIA"].astype(str).isin(cia_sel)
    return df[m].copy()

df_master_filtrado = aplicar_filtros_globais(df_master)
df_div_op_filtrado = aplicar_filtros_globais(df_div_op)
df_sem_div_filtrado = aplicar_filtros_globais(df_sem_div)
df_acao_filtrado = aplicar_filtros_globais(df_acao_total)

mascara_back_m = df_master[COL_GERENTE].astype(str).str.lower().str.contains("suporte backoffice|suporte benner|katia martins", na=False)
mascara_back_d = df_div_op[COL_GERENTE].astype(str).str.lower().str.contains("suporte backoffice|suporte benner|katia martins", na=False)

df_backoffice_dinamico = pd.concat([
    df_backoffice,
    df_master[mascara_back_m],
    df_div_op[mascara_back_d]
], ignore_index=True).drop_duplicates(subset=["Bilhetes"], keep="last") if "Bilhetes" in df_master.columns else df_backoffice

df_backoffice_filtrado = aplicar_filtros_globais(df_backoffice_dinamico)

abas_nomes = [
    "📊 Dashboard & KPIs",
    "🎯 Tratativa Operacional (Geral)",
    "⚠️ Divergência Operação (CIAs/HOT + Tratativa Direta)",
    "✅ Sem Divergência (Conciliação)",
    "🎧 Suporte Backoffice",
    "⚖️ Réplica da Auditoria",
    "📋 Visão Geral da Base Total"
]

if st.session_state["perfil_atual"] == "Compliance":
    abas_nomes.insert(6, "📜 Trilha de Auditoria (Exclusivo Compliance)")
    abas_nomes.insert(7, "⚙️ Gestão de Acessos & Aprovações")

abas_objetos = st.tabs(abas_nomes)

gerentes_base_unicos = sorted([g for g in df_acao_total[COL_GERENTE].dropna().astype(str).unique() if str(g).lower() not in ["suporte backoffice", "suporte benner", "não atribuído", "-", "katia martins"]])

dt_str_export = datetime.datetime.now().strftime("%Y%m%d_%H%M")

# ABA 0: DASHBOARD INTERATIVO - REGRAS DE SOMA E EXCLUSÃO DO TOTAL
with abas_objetos[0]:
    st.subheader("📊 Painel Executivo e Métricas Globais (100% da Base Auditada)")
    
    # REGRA DE CÁLCULO ESTREITA: SOMA APENAS A VISTA, A CREDITO E TAXA
    a_vista_sum = pd.to_numeric(df_acao_filtrado['A vista'], errors='coerce').fillna(0.0).sum() if 'A vista' in df_acao_filtrado.columns else 0.0
    a_credito_sum = pd.to_numeric(df_acao_filtrado['A credito'], errors='coerce').fillna(0.0).sum() if 'A credito' in df_acao_filtrado.columns else 0.0
    val_tarifa = a_vista_sum + a_credito_sum
    
    val_taxa = pd.to_numeric(df_acao_filtrado['Taxa'], errors='coerce').fillna(0.0).sum() if 'Taxa' in df_acao_filtrado.columns else 0.0
    val_receita = pd.to_numeric(df_acao_filtrado['Incentivo'], errors='coerce').fillna(0.0).sum() if 'Incentivo' in df_acao_filtrado.columns else 0.0

    # KPIs Principais
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Bilhetes Auditados em Ação", f"{len(df_acao_filtrado):,}")
    k2.metric("Tarifa Pendente (A vista + A credito)", f"R$ {val_tarifa:,.2f}")
    k3.metric("Taxas Pendentes (Taxa)", f"R$ {val_taxa:,.2f}")
    k4.metric("Receita em Risco (Incentivo)", f"R$ {val_receita:,.2f}")
    
    st.markdown("---")
    
    # PAINEL GRÁFICO 1: DISTRIBUIÇÃO DE INCONSISTÊNCIAS E GERENTES
    col_chart1, col_chart2 = st.columns(2)
    
    with col_chart1:
        st.markdown("##### ⚠️ Classificação Detalhada dos Tipos de Inconsistência")
        if not df_acao_filtrado.empty and "Tipo_Inconsistencia" in df_acao_filtrado.columns:
            df_inc = df_acao_filtrado["Tipo_Inconsistencia"].value_counts().reset_index()
            df_inc.columns = ["Inconsistencia", "Quantidade"]
            
            fig_inc = px.bar(
                df_inc, 
                x="Quantidade", 
                y="Inconsistencia", 
                orientation="h",
                text="Quantidade",
                color="Inconsistencia",
                color_discrete_sequence=["#002060", "#2b9348", "#c1121f", "#7209b7", "#4361ee", "#4cc9f0"]
            )
            fig_inc.update_traces(
                texttemplate='%{text}', 
                textposition='outside',
                cliponaxis=False
            )
            fig_inc.update_layout(
                height=380,
                xaxis_title="Qtd. Bilhetes",
                yaxis_title="",
                showlegend=False,
                margin=dict(l=10, r=30, t=20, b=20)
            )
            st.plotly_chart(fig_inc, use_container_width=True)
        else:
            st.info("Sem dados para exibir o gráfico.")
            
    with col_chart2:
        st.markdown("##### 👤 Volumetria Total por Gerente Responsável")
        if not df_acao_filtrado.empty and COL_GERENTE in df_acao_filtrado.columns:
            df_ger = df_acao_filtrado[COL_GERENTE].value_counts().head(10).reset_index()
            df_ger.columns = ["Gerente", "Quantidade"]
            
            fig_bar = px.bar(
                df_ger, 
                x="Gerente", 
                y="Quantidade", 
                text="Quantidade",
                color="Quantidade",
                color_continuous_scale=["#00509d", "#002060"]
            )
            fig_bar.update_traces(
                texttemplate='%{text}', 
                textposition='outside',
                cliponaxis=False
            )
            fig_bar.update_layout(
                height=380,
                xaxis_title="",
                yaxis_title="Qtd. Bilhetes",
                coloraxis_showscale=False,
                margin=dict(l=10, r=10, t=20, b=20),
                xaxis=dict(tickangle=-25)
            )
            st.plotly_chart(fig_bar, use_container_width=True)
        else:
            st.info("Sem dados para exibir o gráfico.")

    st.markdown("---")
    
    # PAINEL GRÁFICO 2: COMPARATIVO MENSAL POR SETOR
    st.markdown("##### 📈 Volumetria Mensal de Pendências por Setor Responsável")
    if not df_acao_filtrado.empty and "Mes_Ano_Label" in df_acao_filtrado.columns:
        df_mes_setor = df_acao_filtrado.groupby(["Mes_Ano_Label", COL_SETOR]).size().reset_index(name="Quantidade")
        
        fig_line = px.bar(
            df_mes_setor, 
            x="Mes_Ano_Label", 
            y="Quantidade", 
            color=COL_SETOR,
            barmode="group",
            text="Quantidade",
            color_discrete_sequence=["#002060", "#0077b6", "#7209b7", "#4361ee", "#4cc9f0", "#f72585"]
        )
        fig_line.update_traces(textposition='outside', cliponaxis=False)
        fig_line.update_layout(
            height=350,
            xaxis_title="Mês da Emissão",
            yaxis_title="Volume de Bilhetes",
            legend_title="Setor",
            margin=dict(l=10, r=10, t=20, b=20)
        )
        st.plotly_chart(fig_line, use_container_width=True)

# ABA 1: TRATATIVA OPERACIONAL GERAL
with abas_objetos[1]:
    st.subheader("📝 Módulo de Resolução e Detalhamento Operacional")
    
    if len(df_master_filtrado) == 0:
        st.warning("Nenhum bilhete encontrado na base geral para os filtros selecionados.")
    else:
        lista_busca_geral = df_master_filtrado["Bilhetes"].astype(str).tolist()
        opcao_sel_m = st.selectbox("Procure ou Selecione o Bilhete / LOC para Tratativa:", options=lista_busca_geral, key="sb_geral")
        
        row_m = df_master_filtrado[df_master_filtrado["Bilhetes"].astype(str) == opcao_sel_m].iloc[0]
        
        st.markdown(f"""
            <div style="background-color: #fff3cd; border-left: 6px solid #ffc107; padding: 15px; border-radius: 8px; margin: 15px 0;">
                <h4 style="margin: 0; color: #856404;">⚠️ Diagnóstico: Bilhete/LOC {row_m['Bilhetes']}</h4>
                <p style="margin: 5px 0 0 0; color: #856404;">
                    <b>Área / Gerente Resp.:</b> {row_m.get(COL_GERENTE, '-')} | 
                    <b>CIA:</b> {row_m.get('CIA', '-')} | 
                    <b>Status Geral:</b> {row_m.get('Status_Geral', 'Pendente')}
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown("##### 📋 Detalhes do Bilhete e Valores de Emissão")
        cols_det = ["Ponto de venda", "Área Resp. Operação", "CIA", "Bilhetes", "Localizador_Sistema", "Status_Sistema", "Data Emissão", "Pagto", "A vista", "A credito", "Taxa", "Comissão", "Taxa DU", "Desc.", "Incentivo", "VL. Líquido", "Status_Geral", "Obs. Operação"]
        cols_presentes = [c for c in cols_det if c in row_m.index]
        st.dataframe(pd.DataFrame([row_m[cols_presentes]]), hide_index=True)
        st.markdown("---")

        with st.form("form_tratativa_geral"):
            col_a, col_b, col_c, col_d = st.columns([1.2, 1.2, 1.5, 1.2])
            
            with col_a:
                novo_status = st.selectbox("Status da Tratativa:", options=["Já Lançado no ERP", "Pendente de Lançamento", "Aguardando TI", "Cancelado / Devolvido"])
            
            with col_b:
                areas_opcoes = ["Operação", "Suporte backoffice", "Central de Eventos", "Concierge/Lazer", "Unique", "Private"]
                area_atual = str(row_m.get(COL_GERENTE, "Operação"))
                
                if area_atual in ["Silvana Celani", "Silvana Celane"]: area_atual = "Private"
                elif area_atual in ["Jaime Schinaider", "Jaime Schnaider"]: area_atual = "Unique"
                elif area_atual in ["Fabiano Souza"]: area_atual = "Concierge/Lazer"
                elif area_atual in ["Alexandre Souza", "Central de Eventos"]: area_atual = "Central de Eventos"
                elif "suporte" in area_atual.lower() or "benner" in area_atual.lower() or "katia" in area_atual.lower(): area_atual = "Suporte backoffice"
                elif area_atual not in areas_opcoes: area_atual = "Operação"

                index_area = areas_opcoes.index(area_atual)
                nova_area = st.selectbox("Área Responsável (Reatribuir):", options=areas_opcoes, index=index_area)
            
            with col_c:
                if nova_area == "Private":
                    st.text_input("Gerente Responsável:", value="Silvana Celani", disabled=True)
                    gerente_indicado_sel = "Silvana Celani"
                    novo_gerente_texto = ""
                elif nova_area == "Unique":
                    st.text_input("Gerente Responsável:", value="Jaime Schnaider", disabled=True)
                    gerente_indicado_sel = "Jaime Schnaider"
                    novo_gerente_texto = ""
                elif nova_area == "Concierge/Lazer":
                    st.text_input("Gerente Responsável:", value="Fabiano Souza", disabled=True)
                    gerente_indicado_sel = "Fabiano Souza"
                    novo_gerente_texto = ""
                elif nova_area == "Central de Eventos":
                    st.text_input("Gerente Responsável:", value="Alexandre Souza", disabled=True)
                    gerente_indicado_sel = "Alexandre Souza"
                    novo_gerente_texto = ""
                elif nova_area == "Suporte backoffice":
                    st.text_input("Gerente Responsável:", value="Suporte Backoffice", disabled=True)
                    gerente_indicado_sel = "Suporte Backoffice"
                    novo_gerente_texto = ""
                else:
                    gerentes_reservados = ["Silvana Celani", "Jaime Schnaider", "Fabiano Souza", "Alexandre Souza", "Central de Eventos", "Suporte Backoffice", "Katia Martins"]
                    gerentes_operacao_puros = sorted(list(set([g for g in gerentes_base_unicos + ["Keli Santi", "Guilherme Silva", "Ivanete Bertasol", "Rosângela Pallu"] if g not in gerentes_reservados])))
                    
                    if "Outro Gerente..." not in gerentes_operacao_puros:
                        gerentes_operacao_puros.append("Outro Gerente...")
                        
                    gerente_atual_m = str(row_m.get(COL_GERENTE, ""))
                    idx_g = gerentes_operacao_puros.index(gerente_atual_m) if gerente_atual_m in gerentes_operacao_puros else 0
                    
                    gerente_indicado_sel = st.selectbox("Gerente Responsável:", options=gerentes_operacao_puros, index=idx_g)
                    if gerente_indicado_sel == "Outro Gerente...":
                        novo_gerente_texto = st.text_input("Escreva o Nome do Novo Gerente:", placeholder="Digite o nome completo do gerente...")
                    else:
                        novo_gerente_texto = ""

            with col_d:
                num_chamado = st.text_input("Nº do Chamado / Ticket (Obrigatório se Suporte Backoffice):", placeholder="Ex: INC-98472")
                
            obs_detalhe = st.text_area("Observações e Detalhes da Solução:", value="", placeholder="Digite aqui as observações desta tratativa...")
            btn_salvar_g = st.form_submit_button("💾 Salvar Tratativa Operacional")
            
            if btn_salvar_g:
                if nova_area == "Operação":
                    gerente_final = novo_gerente_texto.strip() if gerente_indicado_sel == "Outro Gerente..." else gerente_indicado_sel
                else:
                    gerente_final = gerente_indicado_sel

                if nova_area == "Suporte backoffice" and not num_chamado.strip():
                    st.error("⚠️ Para reatribuir ao **Suporte backoffice**, é OBRIGATÓRIO informar o Número do Chamado!")
                elif nova_area == "Operação" and not gerente_final:
                    st.error("⚠️ Por favor, informe o Nome do Gerente no campo de texto fornecido!")
                else:
                    texto_obs_final = f"[Chamado: {num_chamado}] {obs_detalhe}" if num_chamado else obs_detalhe
                    
                    mascara_m_bilhete = df_master["Bilhetes"].astype(str).str.strip() == str(opcao_sel_m).strip()
                    idx = df_master[mascara_m_bilhete].index
                    
                    novo_log = pd.DataFrame([{
                        "Data_Hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Bilhete": opcao_sel_m,
                        "Usuario_Acao": usuario_log_formatado,
                        "Status_Anterior": row_m.get("Status_Geral", "Pendente"),
                        "Novo_Status": novo_status,
                        "Area_Anterior": row_m.get(COL_GERENTE, "Operação"),
                        "Nova_Area": gerente_final,
                        "Observacao": obs_detalhe,
                        "Tipo_Interacao": "Tratativa Geral"
                    }])

                    if novo_status == "Já Lançado no ERP":
                        rows_upd = df_master.loc[idx].copy()
                        rows_upd["Status_Geral"] = "Já Lançado no ERP"
                        rows_upd[COL_GERENTE] = gerente_final
                        rows_upd["Obs. Operação"] = texto_obs_final
                        rows_upd["Status_Divergencia"] = "Valores Corretos"
                        rows_upd["Tipo_Inconsistencia"] = "Sem Divergência (Conciliado)"
                        
                        df_master = df_master.drop(idx)
                        df_sem_div = pd.concat([df_sem_div, rows_upd], ignore_index=True)
                        msg_res = "transferido para a aba 'Sem Divergência'"
                    else:
                        df_master.loc[idx, "Status_Geral"] = novo_status
                        df_master.loc[idx, COL_GERENTE] = gerente_final
                        df_master.loc[idx, "Obs. Operação"] = texto_obs_final
                        msg_res = "atualizado com sucesso"

                    mascara_back_upd = df_master[COL_GERENTE].astype(str).str.lower().str.contains("suporte backoffice|suporte benner|katia martins", na=False)
                    df_back_atualizado = df_master[mascara_back_upd].copy()
                    df_log_updated = pd.concat([df_log_master, novo_log], ignore_index=True)
                    
                    try:
                        with pd.ExcelWriter(ARQUIVO_DASHBOARD, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            df_master.to_excel(writer, sheet_name="99_Base_Divergencias_Geral", index=False)
                            df_sem_div.to_excel(writer, sheet_name="98_OK_Sem_Divergencia_Concil", index=False)
                            if not df_back_atualizado.empty:
                                df_back_atualizado.to_excel(writer, sheet_name="99_Suporte backoffice", index=False)
                            df_log_updated.to_excel(writer, sheet_name="00_Log_Auditoria", index=False)
                        
                        sincronizar_planilhas_auxiliares(opcao_sel_m, gerente_final, texto_obs_final)
                        
                        st.session_state["msg_sucesso"] = f"✅ Bilhete {opcao_sel_m} {msg_res} por {usuario_log_formatado}! (Atribuído a: {gerente_final})"
                        st.cache_data.clear()
                        st.rerun()
                    except PermissionError:
                        st.error("❌ O arquivo Excel está aberto em outro programa. Feche a planilha para salvar.")

        st.markdown("---")
        col_t1, col_e1 = st.columns([3, 1])
        with col_t1:
            st.markdown(f"### 📊 Lista Completa dos Casos Filtrados ({len(df_master_filtrado)} registros)")
        with col_e1:
            st.download_button(
                label="📥 Exportar Excel Formatado",
                data=gerar_excel_formatado(df_master_filtrado, "Base_Geral"),
                file_name=f"Relatorio_Base_Geral_Filtrado_{dt_str_export}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_exp_geral"
            )
        st.dataframe(df_master_filtrado, hide_index=True)

# ABA 2: DIVERGÊNCIA OPERAÇÃO
with abas_objetos[2]:
    st.subheader("⚠️ Base 98 - Divergência de Operação / CIAs Aéreas / Arquivos HOT")
    
    if len(df_div_op_filtrado) == 0:
        st.warning("Nenhuma divergência de operação encontrada para os filtros selecionados.")
    else:
        lista_busca_div = df_div_op_filtrado["Bilhetes"].astype(str).tolist()
        bilhete_div_sel = st.selectbox("Selecione o Bilhete / LOC para Tratativa Direta:", options=lista_busca_div, key="sb_div_op")
        row_d = df_div_op_filtrado[df_div_op_filtrado["Bilhetes"].astype(str) == bilhete_div_sel].iloc[0]
        
        with st.form("form_tratativa_div_op"):
            c_x, c_y, c_z = st.columns(3)
            with c_x:
                novo_status_d = st.selectbox("Status da Tratativa:", options=["CIA Aérea Corrigida", "Já Lançado no ERP", "Pendente de Lançamento", "Aguardando TI", "Cancelado / Devolvido"])
            with c_y:
                cia_corrigida_d = st.text_input("CIA Aérea Corrigida:", value=str(row_d.get("CIA", "")))
            with c_z:
                num_chamado_d = st.text_input("Nº do Chamado / Ticket (se houver):")
                
            obs_d = st.text_area("Observações e Justificativas:", value="", placeholder="Digite aqui a justificativa ou correção realizada...")
            btn_salvar_d = st.form_submit_button("💾 Salvar Ação nesta Divergência")
            
            if btn_salvar_d:
                mascara_d_bilhete = df_div_op["Bilhetes"].astype(str).str.strip() == str(bilhete_div_sel).strip()
                idx_d = df_div_op[mascara_d_bilhete].index
                texto_obs_d = f"[Correção CIA: {cia_corrigida_d}] " + (f"[Chamado: {num_chamado_d}] " if num_chamado_d else "") + obs_d
                gerente_d_atual = row_d.get(COL_GERENTE, "Operação")
                
                novo_log_d = pd.DataFrame([{
                    "Data_Hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Bilhete": bilhete_div_sel,
                    "Usuario_Acao": usuario_log_formatado,
                    "Status_Anterior": row_d.get("Status_Geral", "Pendente"),
                    "Novo_Status": novo_status_d,
                    "Area_Anterior": gerente_d_atual,
                    "Nova_Area": gerente_d_atual,
                    "Observacao": texto_obs_d,
                    "Tipo_Interacao": "Tratativa Divergência CIA"
                }])

                if novo_status_d in ["Já Lançado no ERP", "CIA Aérea Corrigida"]:
                    rows_upd_d = df_div_op.loc[idx_d].copy()
                    rows_upd_d["Status_Geral"] = novo_status_d
                    if cia_corrigida_d: rows_upd_d["CIA"] = cia_corrigida_d
                    rows_upd_d["Obs. Operação"] = texto_obs_d
                    rows_upd_d["Status_Divergencia"] = "Valores Corretos"
                    rows_upd_d["Tipo_Inconsistencia"] = "Sem Divergência (Conciliado)"
                    
                    df_div_op = df_div_op.drop(idx_d)
                    df_sem_div = pd.concat([df_sem_div, rows_upd_d], ignore_index=True)
                    msg_res_d = "resolvida e transferida para 'Sem Divergência'"
                else:
                    df_div_op.loc[idx_d, "Status_Geral"] = novo_status_d
                    if cia_corrigida_d: df_div_op.loc[idx_d, "CIA"] = cia_corrigida_d
                    df_div_op.loc[idx_d, "Obs. Operação"] = texto_obs_d
                    msg_res_d = "atualizada com sucesso"

                df_log_updated = pd.concat([df_log_master, novo_log_d], ignore_index=True)
                
                try:
                    with pd.ExcelWriter(ARQUIVO_DASHBOARD, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_div_op.to_excel(writer, sheet_name="98_OK_Divergencia_Operacao", index=False)
                        df_sem_div.to_excel(writer, sheet_name="98_OK_Sem_Divergencia_Concil", index=False)
                        df_log_updated.to_excel(writer, sheet_name="00_Log_Auditoria", index=False)
                    
                    sincronizar_planilhas_auxiliares(bilhete_div_sel, gerente_d_atual, texto_obs_d)
                    
                    st.session_state["msg_sucesso"] = f"✅ Divergência {bilhete_div_sel} {msg_res_d} por {usuario_log_formatado}!"
                    st.cache_data.clear()
                    st.rerun()
                except PermissionError:
                    st.error("❌ O arquivo Excel está aberto em outro programa.")
                    
        st.markdown("---")
        col_t2, col_e2 = st.columns([3, 1])
        with col_t2:
            st.markdown(f"### 📊 Lista Completa das Divergências ({len(df_div_op_filtrado)} registros)")
        with col_e2:
            st.download_button(
                label="📥 Exportar Excel Formatado",
                data=gerar_excel_formatado(df_div_op_filtrado, "Divergencias_Operacao"),
                file_name=f"Relatorio_Divergencias_Operacao_Filtrado_{dt_str_export}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_exp_div"
            )
        st.dataframe(df_div_op_filtrado, hide_index=True)

# ABA 3: SEM DIVERGÊNCIA
with abas_objetos[3]:
    st.subheader("✅ Base 98 - Bilhetes Prontos para Conciliação Operacional")
    
    if st.session_state.get("perfil_atual") == "Compliance":
        with st.expander("🛡️ Módulo do Compliance — Retirar Linha e Devolver para Tratativa Operacional", expanded=False):
            if df_sem_div_filtrado.empty or "Bilhetes" not in df_sem_div_filtrado.columns:
                st.info("Nenhum bilhete disponível para devolução nesta visualização.")
            else:
                lista_bilhetes_sd = df_sem_div_filtrado["Bilhetes"].astype(str).tolist()
                bilhete_devolver_sel = st.selectbox("Selecione o Bilhete / LOC para Devolução:", options=lista_bilhetes_sd, key="sb_devolucao_compliance")
                
                row_sd = df_sem_div_filtrado[df_sem_div_filtrado["Bilhetes"].astype(str) == bilhete_devolver_sel].iloc[0]
                
                with st.form("form_devolucao_compliance"):
                    st.write(f"**Bilhete:** {row_sd.get('Bilhetes', '-')} | **Gerente Atual:** {row_sd.get(COL_GERENTE, '-')} | **CIA:** {row_sd.get('CIA', '-')}")
                    
                    area_devolucao_comp = st.selectbox("Devolver para Área:", options=["Operação", "Suporte backoffice", "Central de Eventos", "Concierge/Lazer", "Unique", "Private"])
                    motivo_devolucao = st.text_area("Motivo/Justificativa da Devolução (Obrigatório):", placeholder="Informe o motivo técnico da recusa pelo Compliance...")
                    
                    btn_devolver_sd = st.form_submit_button("🔄 Confirmar Devolução para Operação")
                    
                    if btn_devolver_sd:
                        if not motivo_devolucao.strip():
                            st.error("⚠️ É obrigatório preencher a justificativa da devolução!")
                        else:
                            mascara_sd_bilhete = df_sem_div["Bilhetes"].astype(str).str.strip() == str(bilhete_devolver_sel).strip()
                            idx_sd = df_sem_div[mascara_sd_bilhete].index
                            texto_obs_comp = f"[Devolvido pelo Compliance]: {motivo_devolucao}"
                            
                            area_final_comp = "Suporte Backoffice" if area_devolucao_comp == "Suporte backoffice" else area_devolucao_comp
                            
                            rows_para_devolver = df_sem_div.loc[idx_sd].copy()
                            rows_para_devolver["Status_Geral"] = "Pendente de Lançamento"
                            rows_para_devolver[COL_GERENTE] = area_final_comp
                            rows_para_devolver["Obs. Operação"] = texto_obs_comp
                            rows_para_devolver["Tipo_Inconsistencia"] = "Pendente de Lançamento no ERP"
                            
                            df_sem_div = df_sem_div.drop(idx_sd)
                            df_master = pd.concat([df_master, rows_para_devolver], ignore_index=True)
                            
                            novo_log_comp = pd.DataFrame([{
                                "Data_Hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "Bilhete": bilhete_devolver_sel,
                                "Usuario_Acao": usuario_log_formatado,
                                "Status_Anterior": "Já Lançado no ERP (Sem Divergência)",
                                "Novo_Status": "Pendente de Lançamento (Devolvido pelo Compliance)",
                                "Area_Anterior": row_sd.get(COL_GERENTE, "-"),
                                "Nova_Area": area_final_comp,
                                "Observacao": motivo_devolucao,
                                "Tipo_Interacao": "Devolução do Compliance"
                            }])
                            
                            df_log_updated = pd.concat([df_log_master, novo_log_comp], ignore_index=True)
                            
                            try:
                                with pd.ExcelWriter(ARQUIVO_DASHBOARD, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                    df_master.to_excel(writer, sheet_name="99_Base_Divergencias_Geral", index=False)
                                    df_sem_div.to_excel(writer, sheet_name="98_OK_Sem_Divergencia_Concil", index=False)
                                    df_log_updated.to_excel(writer, sheet_name="00_Log_Auditoria", index=False)
                                
                                sincronizar_planilhas_auxiliares(bilhete_devolver_sel, area_final_comp, texto_obs_comp)
                                
                                st.session_state["msg_sucesso"] = f"🔄 Bilhete {bilhete_devolver_sel} removido da conciliação e devolvido para {area_final_comp}!"
                                st.cache_data.clear()
                                st.rerun()
                            except PermissionError:
                                st.error("❌ O arquivo Excel está aberto em outro programa. Feche a planilha para salvar.")

    col_t3, col_e3 = st.columns([3, 1])
    with col_t3:
        st.markdown(f"### 📊 Lista Completa dos Bilhetes Sem Divergência ({len(df_sem_div_filtrado)} registros)")
    with col_e3:
        st.download_button(
            label="📥 Exportar Excel Formatado",
            data=gerar_excel_formatado(df_sem_div_filtrado, "Sem_Divergencia"),
            file_name=f"Relatorio_Sem_Divergencia_Filtrado_{dt_str_export}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_exp_sem_div"
        )
    st.dataframe(df_sem_div_filtrado, hide_index=True)

# ABA 4: SUPORTE BACKOFFICE
with abas_objetos[4]:
    st.subheader("🎧 Base 99 - Chamados Atribuídos ao Suporte Backoffice / Suporte Benner")
    
    if len(df_backoffice_filtrado) == 0:
        st.info("Nenhum chamado pendente no Suporte Backoffice no momento.")
    else:
        lista_back_bilhetes = df_backoffice_filtrado["Bilhetes"].astype(str).tolist()
        bilhete_back_sel = st.selectbox("Selecione o Chamado / Bilhete para Tratativa do Suporte:", options=lista_back_bilhetes, key="sb_backoffice")
        row_back = df_backoffice_filtrado[df_backoffice_filtrado["Bilhetes"].astype(str) == bilhete_back_sel].iloc[0]

        with st.form("form_solucao_backoffice"):
            acao_back = st.selectbox("Ação do Suporte / Auditoria:", options=["Informar que está Correto (Mover para Sem Divergência)", "Devolver para Tratativa Operacional"])
            
            if acao_back.startswith("Devolver"):
                area_devolucao_bk = st.selectbox("Área Operacional de Destino:", options=["Operação", "Central de Eventos", "Concierge/Lazer", "Unique", "Private"])
            else:
                area_original = str(row_back.get(COL_GERENTE, "Suporte Backoffice"))
                if area_original.lower() in ["suporte backoffice", "suporte benner", "katia martins"]: area_original = "Suporte Backoffice"
                st.text_input("Área Responsável Atribuída:", value=area_original, disabled=True)
                area_devolucao_bk = area_original

            obs_back = st.text_area("Parecer do Suporte Backoffice / Auditoria:", value="", placeholder="Digite aqui o parecer técnico do suporte...")
            btn_salvar_back = st.form_submit_button("💾 Salvar Resolução do Suporte")

            if btn_salvar_back:
                mascara_bk_bilhete = df_master["Bilhetes"].astype(str).str.strip() == str(bilhete_back_sel).strip() if "Bilhetes" in df_master.columns else []
                idx_m_bk = df_master[mascara_bk_bilhete].index if len(mascara_bk_bilhete) > 0 else []
                
                if acao_back.startswith("Informar"):
                    row_upd_bk = row_back.copy()
                    row_upd_bk["Status_Geral"] = "Já Lançado no ERP"
                    row_upd_bk[COL_GERENTE] = area_devolucao_bk
                    row_upd_bk["Status_Divergencia"] = "Valores Corretos"
                    row_upd_bk["Tipo_Inconsistencia"] = "Sem Divergência (Conciliado)"
                    texto_obs_bk = f"[Correto pelo Suporte]: {obs_back}"
                    row_upd_bk["Obs. Operação"] = texto_obs_bk
                    
                    if len(idx_m_bk) > 0: df_master = df_master.drop(idx_m_bk)
                    df_sem_div = pd.concat([df_sem_div, pd.DataFrame([row_upd_bk])], ignore_index=True)
                    
                    novo_status_log = "Já Lançado no ERP (Sem Divergência)"
                    area_destino_log = area_devolucao_bk
                    msg_sucesso_bk = f"🎉 Chamado {bilhete_back_sel} resolvido com área '{area_devolucao_bk}' e movido para **Sem Divergência**!"
                else:
                    texto_obs_bk = f"[Devolvido pelo Suporte]: {obs_back}"
                    if len(idx_m_bk) > 0:
                        df_master.loc[idx_m_bk, "Status_Geral"] = "Pendente de Lançamento"
                        df_master.loc[idx_m_bk, COL_GERENTE] = area_devolucao_bk
                        df_master.loc[idx_m_bk, "Obs. Operação"] = texto_obs_bk
                        
                    novo_status_log = f"Devolvido para {area_devolucao_bk}"
                    area_destino_log = area_devolucao_bk
                    msg_sucesso_bk = f"🔄 Chamado {bilhete_back_sel} devolvido com sucesso para **{area_devolucao_bk}**!"

                novo_log_bk = pd.DataFrame([{
                    "Data_Hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Bilhete": bilhete_back_sel,
                    "Usuario_Acao": usuario_log_formatado,
                    "Status_Anterior": row_back.get("Status_Geral", "Pendente"),
                    "Novo_Status": novo_status_log,
                    "Area_Anterior": "Suporte Backoffice",
                    "Nova_Area": area_destino_log,
                    "Observacao": obs_back,
                    "Tipo_Interacao": "Tratativa Suporte Backoffice"
                }])

                mascara_back_upd = df_master[COL_GERENTE].astype(str).str.lower().str.contains("suporte backoffice|suporte benner|katia martins", na=False)
                df_backoffice_atualizado = df_master[mascara_back_upd].copy()
                df_log_updated = pd.concat([df_log_master, novo_log_bk], ignore_index=True)

                try:
                    with pd.ExcelWriter(ARQUIVO_DASHBOARD, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_master.to_excel(writer, sheet_name="99_Base_Divergencias_Geral", index=False)
                        df_sem_div.to_excel(writer, sheet_name="98_OK_Sem_Divergencia_Concil", index=False)
                        if not df_backoffice_atualizado.empty:
                            df_backoffice_atualizado.to_excel(writer, sheet_name="99_Suporte backoffice", index=False)
                        df_log_updated.to_excel(writer, sheet_name="00_Log_Auditoria", index=False)

                    sincronizar_planilhas_auxiliares(bilhete_back_sel, area_destino_log, texto_obs_bk)

                    st.session_state["msg_sucesso"] = msg_sucesso_bk
                    st.cache_data.clear()
                    st.rerun()
                except PermissionError:
                    st.error("❌ O arquivo Excel está aberto em outro programa.")

        st.markdown("---")
        col_t4, col_e4 = st.columns([3, 1])
        with col_t4:
            st.markdown(f"### 📊 Lista Completa dos Chamados no Suporte ({len(df_backoffice_filtrado)} registros)")
        with col_e4:
            st.download_button(
                label="📥 Exportar Excel Formatado",
                data=gerar_excel_formatado(df_backoffice_filtrado, "Suporte_Backoffice"),
                file_name=f"Relatorio_Suporte_Backoffice_Filtrado_{dt_str_export}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_exp_back"
            )
        st.dataframe(df_backoffice_filtrado, hide_index=True)

# ABA 5: RÉPLICA DA AUDITORIA
with abas_objetos[5]:
    st.subheader("⚖️ Módulo de Contestação e Réplica da Auditoria")
    bilhetes_com_tratativa = df_acao_filtrado[df_acao_filtrado["Status_Geral"] != "Pendente de Lançamento"]
    if len(bilhetes_com_tratativa) > 0:
        bilhete_rep = st.selectbox("Selecione o Bilhete para Réplica:", options=bilhetes_com_tratativa["Bilhetes_Str"].tolist())
        row_rep = df_acao_filtrado[df_acao_filtrado["Bilhetes_Str"] == bilhete_rep].iloc[0]
        
        with st.form("form_replica_auditoria"):
            status_auditoria = st.selectbox("Decisão da Auditoria:", options=["Contestado / Recusado", "Aprovado / Conciliado"])
            area_devolucao = st.selectbox("Devolver para Área / Gerente:", options=["Suporte backoffice", "Central de Eventos", "Concierge/Lazer", "Unique", "Private", "Operação"])
            motivo_replica = st.text_area("Justificativa da Auditoria:", value="", placeholder="Digite a justificativa para a réplica...")
            btn_replica = st.form_submit_button("🚨 Enviar Apontamento")
            
            if btn_replica:
                mascara_rep_bilhete = df_master["Bilhetes"].astype(str).str.strip() == str(bilhete_rep).strip() if "Bilhetes" in df_master.columns else []
                idx_m_rep = df_master[mascara_rep_bilhete].index if len(mascara_rep_bilhete) > 0 else []
                texto_obs_rep = f"[Contestado pela Auditoria]: {motivo_replica}"
                area_rep_final = "Suporte Backoffice" if area_devolucao == "Suporte backoffice" else area_devolucao
                
                novo_log_rep = pd.DataFrame([{
                    "Data_Hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Bilhete": bilhete_rep,
                    "Usuario_Acao": usuario_log_formatado,
                    "Status_Anterior": row_rep.get("Status_Geral", "-"),
                    "Novo_Status": f"Contestado ({area_rep_final})",
                    "Area_Anterior": row_rep.get(COL_GERENTE, "-"),
                    "Nova_Area": area_rep_final,
                    "Observacao": motivo_replica,
                    "Tipo_Interacao": "Réplica da Auditoria"
                }])
                
                if len(idx_m_rep) > 0:
                    df_master.loc[idx_m_rep, "Status_Geral"] = f"Contestado ({area_rep_final})"
                    df_master.loc[idx_m_rep, COL_GERENTE] = area_rep_final
                    df_master.loc[idx_m_rep, "Obs. Operação"] = texto_obs_rep

                df_log_updated = pd.concat([df_log_master, novo_log_rep], ignore_index=True)
                
                try:
                    with pd.ExcelWriter(ARQUIVO_DASHBOARD, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_master.to_excel(writer, sheet_name="99_Base_Divergencias_Geral", index=False)
                        df_log_updated.to_excel(writer, sheet_name="00_Log_Auditoria", index=False)
                    
                    sincronizar_planilhas_auxiliares(bilhete_rep, area_rep_final, texto_obs_rep)

                    st.session_state["msg_sucesso"] = f"✅ Contestação registrada por {usuario_log_formatado}!"
                    st.cache_data.clear()
                    st.rerun()
                except PermissionError:
                    st.error("❌ Feche a planilha para salvar.")

        st.markdown("---")
        st.markdown(f"### 📊 Lista Completa dos Casos em Réplica ({len(bilhetes_com_tratativa)} registros)")
        st.dataframe(bilhetes_com_tratativa, hide_index=True)

idx_aba_compliance = 6

if st.session_state["perfil_atual"] == "Compliance":
    with abas_objetos[idx_aba_compliance]:
        st.subheader("📜 Histórico Completo de Alterações e Trilha de Auditoria")
        
        with st.expander("🗑️ Módulo Master de Gerenciamento e Exclusão de Logs"):
            if not df_log_master.empty:
                indices_log = df_log_master.index.tolist()
                log_opcao = st.selectbox(
                    "Selecione o registro de log que deseja EXCLUIR:",
                    options=indices_log,
                    format_func=lambda i: f"Linha {i} | Data: {df_log_master.loc[i, 'Data_Hora']} | Bilhete: {df_log_master.loc[i, 'Bilhete']} | Usuário: {df_log_master.loc[i, 'Usuario_Acao']}"
                )
                if st.button("❌ Excluir Registro de Log Selecionado"):
                    df_log_master = df_log_master.drop(log_opcao).reset_index(drop=True)
                    try:
                        with pd.ExcelWriter(ARQUIVO_DASHBOARD, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            df_log_master.to_excel(writer, sheet_name="00_Log_Auditoria", index=False)
                        st.session_state["msg_sucesso"] = "🗑️ Registro de log excluído com sucesso!"
                        st.cache_data.clear()
                        st.rerun()
                    except PermissionError:
                        st.error("❌ O arquivo Excel está aberto em outro programa.")
            else:
                st.info("Nenhum registro de log para exclusão.")
                
        st.dataframe(df_log_master, hide_index=False)
    idx_aba_compliance += 1

    with abas_objetos[idx_aba_compliance]:
        st.subheader("⚙️ Central de Aprovações de Acesso e Governança")
        
        usuarios_atuais = carregar_usuarios()
        pendentes = {k: v for k, v in usuarios_atuais.items() if v.get("status") == "PENDENTE"}
        
        if not pendentes:
            st.info("🎉 Nenhuma solicitação de acesso pendente no momento.")
        else:
            for u_id, u_info in pendentes.items():
                st.markdown(f"""
                    <div style="background-color: #ffffff; border-left: 5px solid #002060; padding: 15px; border-radius: 8px; margin-bottom: 12px; box-shadow: 0 2px 6px rgba(0,0,0,0.04);">
                        <b>Nome:</b> {u_info['nome']} | <b>Usuário:</b> {u_id} | <b>Perfil Solicitado:</b> {u_info['perfil']}<br>
                        <small style="color: #6c757d;">Data da Solicitação: {u_info.get('data_solicitacao', '-')}</small>
                    </div>
                """, unsafe_allow_html=True)
                
                col_ap, col_rej, _ = st.columns([1, 1, 4])
                with col_ap:
                    if st.button(f"✅ Aprovar {u_id}", key=f"btn_ap_{u_id}"):
                        usuarios_atuais[u_id]["status"] = "APROVADO"
                        salvar_usuarios(usuarios_atuais)
                        st.session_state["msg_sucesso"] = f"✅ Usuário {u_id} aprovado com sucesso!"
                        st.rerun()
                with col_rej:
                    if st.button(f"❌ Rejeitar {u_id}", key=f"btn_rej_{u_id}"):
                        usuarios_atuais[u_id]["status"] = "REJEITADO"
                        salvar_usuarios(usuarios_atuais)
                        st.session_state["msg_sucesso"] = f"🚫 Usuário {u_id} rejeitado!"
                        st.rerun()

    idx_aba_compliance += 1

with abas_objetos[idx_aba_compliance]:
    col_tv, col_ev = st.columns([3, 1])
    with col_tv:
        st.subheader("📋 Visão Geral da Base Total de Divergências")
    with col_ev:
        st.download_button(
            label="📥 Exportar Excel Formatado",
            data=gerar_excel_formatado(df_acao_filtrado, "Base_Total_Filtrada"),
            file_name=f"Relatorio_Base_Total_Filtrado_{dt_str_export}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_exp_total"
        )
    st.dataframe(df_acao_filtrado, hide_index=True)




###Codigo para versionamento no GitHub###
#----------------------------------------------#

#git add app_revenue_assurance.py
#git commit -m "Adiciona filtro por mes e graficos profissionais Plotly com rotulos de dados"
#git push origin main