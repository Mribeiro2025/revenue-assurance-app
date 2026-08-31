import os
import re
import csv
import datetime
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

DIR_ATUAL = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
os.chdir(DIR_ATUAL)

print("="*75)
print(" MOTOR DE AUDITORIA FP&A ADVANCED - REVENUE ASSURANCE ".center(75, "="))
print("="*75)

MAPA_CIAS = {
    'JJ': 'LATAM', 'LA': 'LATAM', 'PZ': 'LATAM', '4C': 'LATAM', 'XL': 'LATAM', '4M': 'LATAM',
    'G3': 'GOL', 'AD': 'AZUL', 'TP': 'TAP', 'CM': 'COPA', 'AA': 'AMERICAN',
    'UA': 'UNITED', 'DL': 'DELTA', 'AF': 'AIR FRANCE', 'KL': 'KLM', 'AR': 'AEROLINEAS',
    'IB': 'IBERIA', 'UX': 'EUROPA', 'BA': 'BRITISH', 'AV': 'AVIANCA', 'O6': 'AVIANCA',
    'AC': 'CANADA', 'AM': 'AEROMEXICO', 'QR': 'QATAR', 'EK': 'EMIRATES', 'TK': 'TURKISH',
    'LH': 'LUFTHANSA', 'LX': 'SWISS', 'ET': 'ETHIOPIAN', 'AT': 'MAROC', 'SA': 'SOUTH AFRICAN',
    'OB': 'BOLIVIANA', 'PY': 'PARAGUAY', 'HR': 'HAHN', 'AZ': 'ITA AIRWAYS', 'JA': 'JETSMART'
}

def clean_num(val):
    if pd.isna(val) or val is None: return 0.0
    s = str(val).replace('[', '').replace(']', '').replace('R$', '').strip()
    if not s or s == '-': return 0.0
    if ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
    elif ',' in s: s = s.replace(',', '.')
    try: return float(s)
    except: return 0.0

def clean_str_strict(val):
    if pd.isna(val) or val is None: return ''
    s = str(val).strip()
    s = re.sub(r'\.0$', '', s)
    return s

def clean_iata(val):
    if pd.isna(val) or val is None: return ''
    return str(val).replace('-', '').replace(' ', '').strip()

def extract_keys(val):
    if pd.isna(val) or val is None: return []
    s = clean_str_strict(val).upper()
    s = s.replace('[', '').replace(']', '').replace('R$', '').strip()
    if not s or s in ('NAN', 'NONE', 'NULL', '0', '-'): return []
    
    keys = set()
    clean_str = re.sub(r'[^A-Z0-9]', '', s)
    if clean_str:
        keys.add(clean_str)
        clean_no_zeros = clean_str.lstrip('0')
        if clean_no_zeros: keys.add(clean_no_zeros)
        if len(clean_str) == 13 and clean_str.isdigit(): keys.add(clean_str[3:])
        if len(clean_str) >= 10 and clean_str.isdigit(): keys.add(clean_str[-10:])
        
    tokens = re.split(r'[\s/\\,-]+', s)
    for tok in tokens:
        t_clean = re.sub(r'[^A-Z0-9]', '', tok)
        if t_clean:
            keys.add(t_clean)
            if t_clean.isdigit():
                keys.add(t_clean.lstrip('0'))
                if len(t_clean) >= 10: keys.add(t_clean[-10:])
                
    return list(keys)

def carregar_tratativas_e_logs_anteriores(out_file):
    if not os.path.exists(out_file):
        return {}, pd.DataFrame()
    
    try:
        xls = pd.ExcelFile(out_file)
        dict_historico = {}
        
        df_log_antigo = pd.read_excel(xls, sheet_name="00_Log_Auditoria") if "00_Log_Auditoria" in xls.sheet_names else pd.DataFrame()
        
        for aba in ["99_Base_Divergencias_Geral", "98_OK_Divergencia_Operacao", "99_Suporte backoffice"]:
            if aba in xls.sheet_names:
                df_temp = pd.read_excel(xls, sheet_name=aba)
                if "Bilhetes" in df_temp.columns:
                    for _, r in df_temp.iterrows():
                        b_key = clean_str_strict(r["Bilhetes"])
                        if b_key and b_key != "nan":
                            dict_historico[b_key] = {
                                "Status_Geral": r.get("Status_Geral"),
                                "Área Resp. Operação": r.get("Área Resp. Operação"),
                                "Obs. Operação": r.get("Obs. Operação"),
                                "Setor": r.get("Setor"),
                                "Obs_Auditoria_Replica": r.get("Obs_Auditoria_Replica")
                            }
        return dict_historico, df_log_antigo
    except Exception as e:
        print(f"⚠️ Aviso ao carregar histórico anterior: {e}")
        return {}, pd.DataFrame()

def executar_auditoria():
    out_file_model = "Dashboard_Revenue_Assurance_Consolidado.xlsx"
    
    dict_historico, df_log_acumulado = carregar_tratativas_e_logs_anteriores(out_file_model)
    if os.path.exists(out_file_model):
        dt_bkp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            shutil.copy2(out_file_model, f"Backup_Dashboard_{dt_bkp}.xlsx")
            print(f"📦 Backup de segurança gerado: Backup_Dashboard_{dt_bkp}.xlsx")
        except Exception as e:
            print(f"⚠️ Aviso ao criar backup: {e}")

    print("\n[1/6] Carregando Relação IATA...")
    caminho_iata = 'Relacao_Iata_2.xlsx' if os.path.exists('Relacao_Iata_2.xlsx') else 'Relacao_Iata.xlsx'
    df_iata = pd.read_excel(caminho_iata)
    iata_dict = {}

    for idx, r in tqdm(df_iata.iterrows(), total=len(df_iata), desc="Indexando IATAs", unit="reg"):
        k = clean_iata(r.get('Código IATA'))
        if k:
            iata_dict[k] = {
                'Nome_IATA_Oficial': clean_str_strict(r.get('Nome IATA')),
                'Gerente_Responsavel': clean_str_strict(r.get('Gerentes'))
            }

    print("\n[2/6] Indexando Extrato OBT Lemontech...")
    caminho_lemon = 'Extrato_Bilhetes_lemontech.xlsx'
    lemon_index = {}

    if os.path.exists(caminho_lemon):
        df_lemon = pd.read_excel(caminho_lemon, engine='openpyxl')
        for idx, r in tqdm(df_lemon.iterrows(), total=len(df_lemon), desc="Indexando Lemontech", unit="reg"):
            cc_desc = f"{clean_str_strict(r.get('Centro de Custo'))} - {clean_str_strict(r.get('Descrição Centro de Custo'))}".strip(" -")
            info_lemon = {
                'Lemon_Solicitante': clean_str_strict(r.get('Solicitante')) or '-',
                'Lemon_Passageiro': clean_str_strict(r.get('Passageiro')) or '-',
                'Lemon_Consultor': clean_str_strict(r.get('Consultor')) or '-',
                'Lemon_Emissor_Reserva': clean_str_strict(r.get('Emissor Reserva')) or '-',
                'Lemon_Centro_Custo': cc_desc if cc_desc else '-',
                'Lemon_Forma_Pagto': clean_str_strict(r.get('Forma de Pagamento')) or '-',
                'Lemon_Autorizacao_Cartao': clean_str_strict(r.get('Autorização do Cartão')) or '-',
                'Lemon_OnOff': clean_str_strict(r.get('On|Off')) or '-',
                'Lemon_Source': clean_str_strict(r.get('Source')) or '-'
            }
            for col_k in ['Bilhete', 'Nº Pedido', 'Solicitação']:
                for k in extract_keys(r.get(col_k)):
                    if k not in lemon_index:
                        lemon_index[k] = info_lemon

    print("\n[3/6] Indexando ERP Benner (Multicolunas)...")
    df_benner = pd.read_excel('Acumulado.xlsx', sheet_name='Planilha1')
    benner_index = {}
    key_cols_benner = ['Bilhete', 'Localizador', 'Rloc Cia', 'Código Rloc', 'Pedido', 'Fatura/Documento', 'Handle Accounting', 'Accounting VM', 'Apurações Fee']

    for idx, r in tqdm(df_benner.iterrows(), total=len(df_benner), desc="Indexando Benner", unit="linha"):
        info = {
            'Benner_Index': idx,
            'Benner_Situação': clean_str_strict(r.get('Situação')) or 'ATIVO',
            'Benner_Fornecedor': clean_str_strict(r.get('Apelido Fornecedor') or r.get('Fornecedor')) or '-',
            'Benner_Localizador': clean_str_strict(r.get('Rloc Cia') or r.get('Localizador') or r.get('Código Rloc')) or '-',
            'Benner_Bilhete': clean_str_strict(r.get('Bilhete')) or '-',
            'Benner_Tarifa': clean_num(r.get('Tarifa')),
            'Benner_Taxas': clean_num(r.get('Taxas')) + clean_num(r.get('Taxa BR')),
            'Benner_Receita': clean_num(r.get('Comissão')) + clean_num(r.get('Taxa DU')) + clean_num(r.get('Incentivo')) + clean_num(r.get('Fee')) + clean_num(r.get('RAV')),
            'Benner_Cliente': clean_str_strict(r.get('Cliente') or r.get('Apelido Cliente')) or '-',
            'Benner_Emissor': clean_str_strict(r.get('Emissor') or r.get('Agente Criação')) or '-',
            'Benner_Passageiro': clean_str_strict(r.get('Passageiro')) or '-',
            'Benner_Sistema_Reserva': clean_str_strict(r.get('Sistema Reserva')) or '-'
        }
        for col in key_cols_benner:
            for k in extract_keys(r.get(col)):
                if k not in benner_index:
                    benner_index[k] = info

    print("\n[4/6] Indexando Relatório Sabre...")
    caminho_sabre = 'Relacao_Sabre_3.xlsx' if os.path.exists('Relacao_Sabre_3.xlsx') else 'Relacao_Sabre.xlsx'
    lines = pd.read_excel(caminho_sabre).iloc[:, 0].dropna().astype(str).tolist()
    sabre_rows = []

    for line in lines:
        if 'Total Amount' in line or 'Total Commission' in line or 'PCC,DATE,AL CODE' in line: continue
        parts = list(csv.reader([line]))[0]
        if len(parts) >= 15 and parts[0] != 'PCC':
            cleaned = [p.replace('="', '').replace('"', '').strip() for p in parts]
            sabre_rows.append(cleaned[:15])

    df_sabre = pd.DataFrame(sabre_rows, columns=['PCC', 'DATE', 'AL_CODE', 'TICKET_NUM', 'PNR', 'NAME', 'LAST_NAME', 'CUR', 'OB', 'COM_AMT', 'TOTAL', 'FOP', 'AGT', 'TIME', 'STATUS'])
    sabre_index = {}

    for idx, r in tqdm(df_sabre.iterrows(), total=len(df_sabre), desc="Indexando Sabre", unit="reg"):
        info = {
            'Sabre_PCC': r['PCC'],
            'Sabre_AGT': r['AGT'],
            'Sabre_Passageiro': f"{r['NAME']} {r['LAST_NAME']}".strip()
        }
        for k in extract_keys(r['TICKET_NUM']) + extract_keys(r['PNR']):
            if k not in sabre_index:
                sabre_index[k] = info

    print("\n[5/6] Auditando e Conciliando Emissões das Cias Aéreas...")
    fontes = [('AZUL.XLSX', 'Azul'), ('BSP.XLSX', 'BSP'), ('HOT.XLSX', 'HOT')]
    registros_conciliados = []

    for caminho_arq, nome_fonte in fontes:
        if not os.path.exists(caminho_arq): continue
        df_raw = pd.read_excel(caminho_arq, header=None)
        curr_iata, curr_ponto_venda = '', ''
        
        for r in tqdm(range(len(df_raw)), desc=f"Auditando {nome_fonte}", unit="linha"):
            col0 = clean_str_strict(df_raw.iloc[r, 0])
            col1 = clean_str_strict(df_raw.iloc[r, 1])
            col3 = clean_str_strict(df_raw.iloc[r, 3])
            col4 = clean_str_strict(df_raw.iloc[r, 4])
            col5 = clean_str_strict(df_raw.iloc[r, 5])
            
            if re.match(r'^\d{2}-\d{5}', col0) or re.match(r'^\d{7,8}$', col0):
                curr_iata = col0
                if col4: curr_ponto_venda = col4
                continue
                
            if col3 != '' and col5 != '':
                if col3.upper() in ['BILHETE\\RLOC', 'BILHETE', 'RLOC'] or col5.upper() in ['EMISSÃO', 'EMISSAO'] or col3 == '[]':
                    continue
                    
                iata_clean = clean_iata(curr_iata)
                m_iata = iata_dict.get(iata_clean, {})
                nome_iata_oficial = m_iata.get('Nome_IATA_Oficial', 'Não Encontrado na Relação IATA')
                gerente_resp = m_iata.get('Gerente_Responsavel', 'Não Mapeado')
                doc_val = clean_str_strict(df_raw.iloc[r, 23])
                
                keys_emissao = set(extract_keys(col3) + extract_keys(doc_val))
                
                b_match = next((benner_index[ek] for ek in keys_emissao if ek in benner_index), None)
                s_match = next((sabre_index[ek] for ek in keys_emissao if ek in sabre_index), None)
                l_match = next((lemon_index[ek] for ek in keys_emissao if ek in lemon_index), None)
                
                a_vista = clean_num(df_raw.iloc[r, 9])
                a_credito = clean_num(df_raw.iloc[r, 11])
                tarifa_emitida = a_vista + a_credito
                taxa_emitida = clean_num(df_raw.iloc[r, 12])
                comissao = clean_num(df_raw.iloc[r, 14])
                taxa_du = clean_num(df_raw.iloc[r, 15])
                incentivo = clean_num(df_raw.iloc[r, 19])
                receita_emitida = comissao + taxa_du + incentivo
                vl_liquido = clean_num(df_raw.iloc[r, 21])
                
                gerente_norm = gerente_resp.strip().upper()
                if 'JAIME SCHNAIDER' in gerente_norm:
                    setor_inicial = 'Unique'
                elif 'CENTRAL DE EVENTOS' in gerente_norm:
                    setor_inicial = 'Central de Eventos'
                elif 'FABIANO SOUZA' in gerente_norm:
                    setor_inicial = 'Concierge/Lazer'
                else:
                    setor_inicial = 'Operação'
                    
                area_op_norm = gerente_resp.strip().upper()
                pv_norm = curr_ponto_venda.strip().upper()
                
                if 'CENTRAL DE EVENTOS' in area_op_norm or 'EVENTO' in area_op_norm or 'EVENTOS' in pv_norm:
                    setor_final = 'Central de Eventos'
                elif 'PRIVATE' in area_op_norm or 'PRIVATE' in pv_norm:
                    setor_final = 'Private'
                elif 'BENNER' in area_op_norm or 'BACKOFFICE' in area_op_norm or 'SUPORTE' in area_op_norm:
                    setor_final = 'Suporte backoffice'
                else:
                    setor_final = setor_inicial

                if b_match:
                    status_sistema = b_match['Benner_Situação']
                    fornec_sistema = b_match['Benner_Fornecedor']
                    loc_sistema = b_match['Benner_Localizador']
                    bilhete_sistema = b_match['Benner_Bilhete']
                    tarifa_sistema = b_match['Benner_Tarifa']
                    taxa_sistema = b_match['Benner_Taxas']
                    receita_sistema = b_match['Benner_Receita']
                    cliente_sistema = b_match['Benner_Cliente']
                    emissor_sistema = b_match['Benner_Emissor']
                    sist_reserva = b_match['Benner_Sistema_Reserva']
                    
                    sigla_cia = col1.strip().upper()
                    nome_esperado_cia = MAPA_CIAS.get(sigla_cia, sigla_cia)
                    status_cia = 'OK' if (nome_esperado_cia in fornec_sistema.upper() or sigla_cia in fornec_sistema.upper()) else 'Divergência de Cia Aérea'
                        
                    dif_tarifa = tarifa_emitida - tarifa_sistema
                    dif_taxa = taxa_emitida - taxa_sistema
                    dif_receita = receita_emitida - receita_sistema
                    
                    if nome_fonte == 'HOT':
                        if status_cia != 'OK':
                            status_divergencia = 'Divergência de Cia Aérea'
                        else:
                            status_divergencia = 'Valores Corretos'
                    else:
                        divs = []
                        if round(abs(dif_tarifa), 2) >= 0.01: divs.append('Tarifa')
                        if round(abs(dif_taxa), 2) >= 0.01: divs.append('Taxa')
                        if round(abs(dif_receita), 2) >= 0.01: divs.append('Receita')
                        
                        if divs: status_divergencia = f"Divergência de {' e '.join(divs)}"
                        elif status_cia != 'OK': status_divergencia = 'Divergência de Cia Aérea'
                        else: status_divergencia = 'Valores Corretos'
                        
                    status_geral = 'Emitido e Lançado'
                    aba_destino = '98_OK_Sem_Divergencia_Concil' if status_divergencia == 'Valores Corretos' else '98_OK_Divergencia_Operacao'
                else:
                    status_sistema = 'NAO_CONSTA'
                    fornec_sistema = '-'
                    loc_sistema = col3
                    bilhete_sistema = col3
                    tarifa_sistema = 0.0
                    taxa_sistema = 0.0
                    receita_sistema = 0.0
                    cliente_sistema = curr_ponto_venda
                    emissor_sistema = '-'
                    sist_reserva = '-'
                    status_cia = 'Pendente'
                    dif_tarifa = tarifa_emitida
                    dif_taxa = taxa_emitida
                    dif_receita = receita_emitida
                    status_divergencia = 'Pendente de Lançamento'
                    status_geral = 'Pendente de Lançamento (Não Consta)'
                    aba_destino = '99_Base_Divergencias_Geral'

                bilhete_chave = clean_str_strict(col3)
                obs_op = 'Sem tratativa na operação'
                obs_replica = '-'
                
                if bilhete_chave in dict_historico:
                    hist_data = dict_historico[bilhete_chave]
                    if hist_data.get("Status_Geral"): status_geral = hist_data["Status_Geral"]
                    if hist_data.get("Área Resp. Operação"): gerente_resp = hist_data["Área Resp. Operação"]
                    if hist_data.get("Obs. Operação"): obs_op = hist_data["Obs. Operação"]
                    if hist_data.get("Setor"): setor_final = hist_data["Setor"]
                    if hist_data.get("Obs_Auditoria_Replica"): obs_replica = hist_data["Obs_Auditoria_Replica"]

                rec = {
                    'Ponto de venda': curr_ponto_venda,
                    'Código Iata': curr_iata,
                    'Gerentes': gerente_resp,
                    'CIA': col1,
                    'Fornecedor_Sistema': fornec_sistema,
                    'Status_Cia': status_cia,
                    'Bilhetes': bilhete_chave,
                    'Localizador_Sistema': loc_sistema,
                    'Status_Sistema': status_sistema,
                    'Data Emissão': col5,
                    'Pagto': clean_str_strict(df_raw.iloc[r, 6]),
                    'A vista': a_vista,
                    'A credito': a_credito,
                    'Tarifa_Sistema': tarifa_sistema,
                    'Dif_Tarifa': dif_tarifa,
                    'Taxa': taxa_emitida,
                    'Taxa_Sistema': taxa_sistema,
                    'Dif_Taxa': dif_taxa,
                    'Comissão': comissao,
                    'Taxa DU': taxa_du,
                    'Desc.': clean_num(df_raw.iloc[r, 16]),
                    'Incentivo': incentivo,
                    'Receita_Sistema': receita_sistema,
                    'Dif_Receita': dif_receita,
                    'VL. Líquido': vl_liquido,
                    'Status_Divergencia': status_divergencia,
                    'Consultor': s_match['Sabre_AGT'] if s_match else (l_match['Lemon_Consultor'] if l_match else '-'),
                    'Status_Geral': status_geral,
                    'Área Resp. Operação': gerente_resp,
                    'Obs. Operação': obs_op,
                    'Obs_Auditoria_Replica': obs_replica,
                    'Emissor': emissor_sistema if b_match else (l_match['Lemon_Emissor_Reserva'] if l_match else '-'),
                    'Sistema Reserva': sist_reserva if b_match else (l_match['Lemon_Source'] if l_match else '-'),
                    'Cliente': cliente_sistema,
                    'Setor': setor_final,
                    'Aba_Destino': aba_destino,
                    'Tipo_Emissao_Lemon': l_match['Lemon_OnOff'] if l_match else '-',
                    'Consultor_Lemon': l_match['Lemon_Consultor'] if l_match else '-',
                    'Emissor_Reserva_Lemon': l_match['Lemon_Emissor_Reserva'] if l_match else '-',
                    'Centro_Custo_Lemon': l_match['Lemon_Centro_Custo'] if l_match else '-',
                    'Forma_Pagto_Lemon': l_match['Lemon_Forma_Pagto'] if l_match else '-',
                    'Autorizacao_Cartao_Lemon': l_match['Lemon_Autorizacao_Cartao'] if l_match else '-'
                }
                registros_conciliados.append(rec)

    df_master = pd.DataFrame(registros_conciliados)

    cols_34 = [
        'Ponto de venda', 'Código Iata', 'Gerentes', 'CIA', 'Fornecedor_Sistema', 'Status_Cia', 
        'Bilhetes', 'Localizador_Sistema', 'Status_Sistema', 'Data Emissão', 'Pagto', 'A vista', 
        'A credito', 'Tarifa_Sistema', 'Dif_Tarifa', 'Taxa', 'Taxa_Sistema', 'Dif_Taxa', 
        'Comissão', 'Taxa DU', 'Desc.', 'Incentivo', 'Receita_Sistema', 'Dif_Receita', 
        'VL. Líquido', 'Status_Divergencia', 'Consultor', 'Status_Geral', 'Área Resp. Operação', 
        'Obs. Operação', 'Emissor', 'Sistema Reserva', 'Cliente', 'Setor'
    ]

    cols_29_furo = [
        'Ponto de venda', 'Código Iata', 'Gerentes', 'CIA', 'Bilhetes', 'Localizador_Sistema', 
        'Status_Sistema', 'Data Emissão', 'Pagto', 'A vista', 'A credito', 'Taxa', 'Comissão', 
        'Taxa DU', 'Desc.', 'Incentivo', 'VL. Líquido', 'Status_Geral', 'Área Resp. Operação', 
        'Obs. Operação', 'Tipo_Emissao_Lemon', 'Consultor_Lemon', 'Emissor_Reserva_Lemon', 
        'Centro_Custo_Lemon', 'Forma_Pagto_Lemon', 'Autorizacao_Cartao_Lemon', 'Sistema Reserva', 'Cliente', 'Setor'
    ]

    pareto_df = df_master.groupby(['Setor', 'Ponto de venda']).agg(
        Qtd_Bilhetes=('Bilhetes', 'count'),
        Tarifa_Pendente_R_=('A credito', 'sum'),
        Taxa_Pendente_R_=('Taxa', 'sum'),
        Receita_Pendente_R_=('Incentivo', 'sum')
    ).reset_index().rename(columns={
        'Ponto de venda': 'Cliente / Ponto de Venda',
        'Tarifa_Pendente_R_': 'Tarifa_Pendente_R$',
        'Taxa_Pendente_R_': 'Taxa_Pendente_R$',
        'Receita_Pendente_R_': 'Receita_Pendente_R$'
    }).sort_values(by=['Qtd_Bilhetes', 'Tarifa_Pendente_R$'], ascending=False)

    df_98_div = df_master[df_master['Aba_Destino'] == '98_OK_Divergencia_Operacao'][cols_34] if len(df_master[df_master['Aba_Destino'] == '98_OK_Divergencia_Operacao']) > 0 else pd.DataFrame(columns=cols_34)
    df_98_ok = df_master[df_master['Aba_Destino'] == '98_OK_Sem_Divergencia_Concil'][cols_34] if len(df_master[df_master['Aba_Destino'] == '98_OK_Sem_Divergencia_Concil']) > 0 else pd.DataFrame(columns=cols_34)
    df_99_gen = df_master[df_master['Aba_Destino'] == '99_Base_Divergencias_Geral'][cols_29_furo] if len(df_master[df_master['Aba_Destino'] == '99_Base_Divergencias_Geral']) > 0 else pd.DataFrame(columns=cols_29_furo)

    df_99_suporte = df_99_gen[df_99_gen['Setor'] == 'Suporte backoffice'] if len(df_99_gen) > 0 else pd.DataFrame(columns=cols_29_furo)
    df_99_eventos = df_99_gen[df_99_gen['Setor'] == 'Central de Eventos'] if len(df_99_gen) > 0 else pd.DataFrame(columns=cols_29_furo)
    df_99_lazer = df_99_gen[df_99_gen['Setor'] == 'Concierge/Lazer'] if len(df_99_gen) > 0 else pd.DataFrame(columns=cols_29_furo)
    df_99_unique = df_99_gen[df_99_gen['Setor'] == 'Unique'] if len(df_99_gen) > 0 else pd.DataFrame(columns=cols_29_furo)
    df_99_private = df_99_gen[df_99_gen['Setor'] == 'Private'] if len(df_99_gen) > 0 else pd.DataFrame(columns=cols_29_furo)
    df_99_operacao = df_99_gen[df_99_gen['Setor'] == 'Operação'] if len(df_99_gen) > 0 else pd.DataFrame(columns=cols_29_furo)

    print("\n[6/6] Exportando Dashboard Formatado em Excel...")

    with pd.ExcelWriter(out_file_model, engine='openpyxl') as writer:
        pareto_df.to_excel(writer, sheet_name='01_Pareto_Cliente', index=False)
        df_98_div.to_excel(writer, sheet_name='98_OK_Divergencia_Operacao', index=False)
        df_98_ok.to_excel(writer, sheet_name='98_OK_Sem_Divergencia_Concil', index=False)
        df_99_gen.to_excel(writer, sheet_name='99_Base_Divergencias_Geral', index=False)
        df_99_suporte.to_excel(writer, sheet_name='99_Suporte backoffice', index=False)
        df_99_eventos.to_excel(writer, sheet_name='99_Central de Eventos', index=False)
        df_99_lazer.to_excel(writer, sheet_name='99_Concierge-Lazer', index=False)
        df_99_unique.to_excel(writer, sheet_name='99_Unique', index=False)
        df_99_private.to_excel(writer, sheet_name='99_Private', index=False)
        df_99_operacao.to_excel(writer, sheet_name='99_Operação', index=False)
        
        if not df_log_acumulado.empty:
            df_log_acumulado.to_excel(writer, sheet_name='00_Log_Auditoria', index=False)
        else:
            pd.DataFrame(columns=[
                "Data_Hora", "Bilhete", "Usuario_Acao", "Status_Anterior", 
                "Novo_Status", "Area_Anterior", "Nova_Area", "Observacao", "Tipo_Interacao"
            ]).to_excel(writer, sheet_name='00_Log_Auditoria', index=False)

    wb = openpyxl.load_workbook(out_file_model)

    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    fill_alerta = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_alerta = Font(color="9C0006", bold=True)

    currency_cols = [
        'Tarifa_Pendente_R$', 'Taxa_Pendente_R$', 'Receita_Pendente_R$', 'Receita_Risco',
        'A vista', 'A credito', 'Tarifa_Sistema', 'Dif_Tarifa', 'Taxa', 'Taxa_Sistema', 
        'Dif_Taxa', 'Comissão', 'Taxa DU', 'Desc.', 'Incentivo', 'Receita_Sistema', 
        'Dif_Receita', 'VL. Líquido'
    ]

    for sheetname in tqdm(wb.sheetnames, desc="Formatando Abas do Excel", unit="aba"):
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True
        
        if sheetname not in ['01_Pareto_Cliente', '00_Log_Auditoria']:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        max_row = ws.max_row
        
        if max_row > 1 and sheetname not in ['01_Pareto_Cliente', '00_Log_Auditoria']:
            ws.cell(row=max_row + 1, column=1, value="TOTAL").font = Font(bold=True)
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                col_name = str(ws.cell(row=1, column=col_idx).value or '')
                if col_name in currency_cols or col_name in ['Qtd_Bilhetes']:
                    c_tot = ws.cell(row=max_row + 1, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{max_row})")
                    c_tot.font = Font(bold=True)

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_name = str(ws.cell(row=1, column=col_idx).value or '')
            
            max_len = max([len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)])
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
            if col_name in currency_cols:
                for r in range(2, max_row + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    if cell.value not in ['-', None]:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = 'R$ #,##0.00'
                        except: pass
                        
            if col_name.startswith('Dif_'):
                for r in range(2, max_row + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    try:
                        if round(abs(float(cell.value)), 2) >= 0.01:
                            cell.fill = fill_alerta
                            cell.font = font_alerta
                    except: pass

    wb.save(out_file_model)

    print("\n" + "="*75)
    print(f" AUDITORIA CONCLUÍDA COM SUCESSO! SALVO EM: {out_file_model} ".center(75, "="))
    print("="*75)

if __name__ == '__main__':
    executar_auditoria()